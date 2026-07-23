import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime, timedelta
import yfinance as yf
import warnings
import io
warnings.filterwarnings("ignore")

# --- CONFIGURATION ---
INDICES = {
    "S&P 500": "^GSPC",
    "NASDAQ 100": "^IXIC",
    "Dow Jones": "^DJI",
    "Russell 2000": "^RUT",
    "VIX": "^VIX"
}

SECTORS = {
    "Matériaux": "XLB",
    "Communication": "XLC",
    "Énergie": "XLE",
    "Finance": "XLF",
    "Industrie": "XLI",
    "Technologie": "XLK",
    "Consommation de base": "XLP",
    "Utilities": "XLU",
    "Santé": "XLV",
    "Consommation discrétionnaire": "XLY",
    "Immobilier": "XLRE"
}

# --- FONCTIONS ---
def fetch_data(symbols, period="5y", interval="1d"):
    df = pd.DataFrame()
    for name, symbol in symbols.items():
        data = yf.download(symbol, period=period, interval=interval)
        if not data.empty:
            df[name] = data['Close']
    return df

def calculate_indicators(df):
    new_df = df.copy()
    for col in df.columns:
        new_df[f"{col}_EMA9"] = df[col].ewm(span=9, adjust=False).mean()
        new_df[f"{col}_SMA20"] = df[col].rolling(window=20).mean()
        new_df[f"{col}_SMA50"] = df[col].rolling(window=50).mean()
        new_df[f"{col}_SMA200"] = df[col].rolling(window=200).mean()
        new_df[f"{col}_Momentum_5d"] = (df[col] / df[col].shift(5)) - 1

        delta = df[col].diff()
        gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
        loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
        rs = gain / loss
        new_df[f"{col}_RSI"] = 100 - (100 / (1 + rs))

        ema12 = df[col].ewm(span=12, adjust=False).mean()
        ema26 = df[col].ewm(span=26, adjust=False).mean()
        new_df[f"{col}_MACD"] = ema12 - ema26
        new_df[f"{col}_MACD_Signal"] = new_df[f"{col}_MACD"].ewm(span=9, adjust=False).mean()
        new_df[f"{col}_MACD_Hist"] = new_df[f"{col}_MACD"] - new_df[f"{col}_MACD_Signal"]

        new_df[f"{col}_BB_Middle"] = df[col].rolling(window=20).mean()
        rolling_std = df[col].rolling(window=20).std()
        new_df[f"{col}_BB_Upper"] = new_df[f"{col}_BB_Middle"] + (rolling_std * 2)
        new_df[f"{col}_BB_Lower"] = new_df[f"{col}_BB_Middle"] - (rolling_std * 2)
    return new_df

def analyze_trends(df):
    latest = df.iloc[-1]
    trends = {}
    for name in df.columns:
        if any(suffix in name for suffix in ["_EMA", "_SMA", "_Momentum", "_RSI", "_MACD", "_BB"]):
            continue
        price = latest[name]
        if pd.isna(price):
            continue

        if name == "S&P 500":
            price *= 10; ema9 = latest.get(f"{name}_EMA9", float('nan')) * 10
            sma20 = latest.get(f"{name}_SMA20", float('nan')) * 10
            sma50 = latest.get(f"{name}_SMA50", float('nan')) * 10
            sma200 = latest.get(f"{name}_SMA200", float('nan')) * 10
            rsi = latest.get(f"{name}_RSI", float('nan'))
            macd = latest.get(f"{name}_MACD", float('nan')) * 10
            macd_signal = latest.get(f"{name}_MACD_Signal", float('nan')) * 10
        elif name == "NASDAQ 100":
            price *= 40; ema9 = latest.get(f"{name}_EMA9", float('nan')) * 40
            sma20 = latest.get(f"{name}_SMA20", float('nan')) * 40
            sma50 = latest.get(f"{name}_SMA50", float('nan')) * 40
            sma200 = latest.get(f"{name}_SMA200", float('nan')) * 40
            rsi = latest.get(f"{name}_RSI", float('nan'))
            macd = latest.get(f"{name}_MACD", float('nan')) * 40
            macd_signal = latest.get(f"{name}_MACD_Signal", float('nan')) * 40
        elif name == "Dow Jones":
            price *= 100; ema9 = latest.get(f"{name}_EMA9", float('nan')) * 100
            sma20 = latest.get(f"{name}_SMA20", float('nan')) * 100
            sma50 = latest.get(f"{name}_SMA50", float('nan')) * 100
            sma200 = latest.get(f"{name}_SMA200", float('nan')) * 100
            rsi = latest.get(f"{name}_RSI", float('nan'))
            macd = latest.get(f"{name}_MACD", float('nan')) * 100
            macd_signal = latest.get(f"{name}_MACD_Signal", float('nan')) * 100
        elif name == "Russell 2000":
            price *= 10; ema9 = latest.get(f"{name}_EMA9", float('nan')) * 10
            sma20 = latest.get(f"{name}_SMA20", float('nan')) * 10
            sma50 = latest.get(f"{name}_SMA50", float('nan')) * 10
            sma200 = latest.get(f"{name}_SMA200", float('nan')) * 10
            rsi = latest.get(f"{name}_RSI", float('nan'))
            macd = latest.get(f"{name}_MACD", float('nan')) * 10
            macd_signal = latest.get(f"{name}_MACD_Signal", float('nan')) * 10
        else:
            ema9 = latest.get(f"{name}_EMA9", float('nan'))
            sma20 = latest.get(f"{name}_SMA20", float('nan'))
            sma50 = latest.get(f"{name}_SMA50", float('nan'))
            sma200 = latest.get(f"{name}_SMA200", float('nan'))
            rsi = latest.get(f"{name}_RSI", float('nan'))
            macd = latest.get(f"{name}_MACD", float('nan'))
            macd_signal = latest.get(f"{name}_MACD_Signal", float('nan'))

        trends[name] = {
            "Prix": round(price, 2),
            "> EMA9": "✅" if not pd.isna(ema9) and price > ema9 else "❌",
            "> SMA20": "✅" if not pd.isna(sma20) and price > sma20 else "❌",
            "> SMA50": "✅" if not pd.isna(sma50) and price > sma50 else "❌",
            "> SMA200": "✅" if not pd.isna(sma200) and price > sma200 else "❌",
            "RSI": f"{round(rsi, 2) if not pd.isna(rsi) else 'N/A'}",
            "RSI_Surachat": "✅" if not pd.isna(rsi) and rsi > 70 else "❌",
            "RSI_Survente": "✅" if not pd.isna(rsi) and rsi < 30 else "❌",
            "MACD_Achat": "✅" if not pd.isna(macd) and not pd.isna(macd_signal) and macd > macd_signal else "❌",
            "BB_Position": f"{(price - (latest.get(f'{name}_BB_Lower', float('nan')))) / ((latest.get(f'{name}_BB_Upper', float('nan'))) - (latest.get(f'{name}_BB_Lower', float('nan')))) * 100:.1f}%" if not pd.isna(latest.get(f"{name}_BB_Upper", float('nan'))) and not pd.isna(latest.get(f"{name}_BB_Lower", float('nan'))) else "N/A"
        }
    return pd.DataFrame(trends).T

def generate_excel(df_indices, trends):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    import os

    os.makedirs("output", exist_ok=True)
    wb = Workbook()

    # Feuille Tendances
    ws_tendances = wb.active
    ws_tendances.title = "Tendances"
    trends_reset = trends.reset_index()
    trends_reset.rename(columns={'index': 'Indice/Secteur'}, inplace=True)
    for c in range(len(trends_reset.columns)):
        ws_tendances.cell(row=1, column=c+1, value=trends_reset.columns[c])
    for r in range(len(trends_reset)):
        for c in range(len(trends_reset.columns)):
            ws_tendances.cell(row=r+2, column=c+1, value=trends_reset.iloc[r, c])

    # Feuille Données
    ws_donnees = wb.create_sheet("Données")
    for r in range(len(df_indices.index) + 1):
        for c in range(len(df_indices.columns) + 1):
            if r == 0:
                ws_donnees.cell(row=r+1, column=c+1, value=df_indices.columns[c-1] if c > 1 else df_indices.index.name)
            else:
                ws_donnees.cell(row=r+1, column=c+1, value=df_indices.iloc[r-1, c-1] if c > 1 else str(df_indices.index[r-1]))

    wb.save("output/rapport_marche.xlsx")
    return "output/rapport_marche.xlsx"

# --- INTERFACE STREAMLIT ---
st.title("📈 Analyse Technique des Marchés")
st.markdown("---")

# Sélecteur de période
period = st.selectbox(
    "Période",
    ["1d", "5d", "1mo", "3mo", "6mo", "1y", "2y", "5y", "max"],
    index=6  # 5y par défaut
)

# Charge les données
with st.spinner("Chargement des données..."):
    df_indices = fetch_data(INDICES, period=period, interval="1d")
    df_sectors = fetch_data(SECTORS, period=period, interval="1d")

    df_indices = calculate_indicators(df_indices)
    df_sectors = calculate_indicators(df_sectors)

    trends_indices = analyze_trends(df_indices)
    trends_sectors = analyze_trends(df_sectors)

# Onglets
tab1, tab2, tab3 = st.tabs(["📊 Tendances", "📈 Graphiques", "🔍 Analyse"])

with tab1:
    st.subheader("Tendances des Indices")
    st.dataframe(trends_indices, width='stretch')

    st.subheader("Tendances des Secteurs")
    st.dataframe(trends_sectors, width='stretch')

    # Bouton pour exporter en Excel
    if st.button("📥 Exporter en Excel"):
        excel_file = generate_excel(df_indices, pd.concat([trends_indices, trends_sectors]))
        with open(excel_file, "rb") as f:
            st.download_button(
                label="Télécharger le rapport Excel",
                data=f,
                file_name="rapport_marche.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

with tab2:
    # Sélecteur: Indices ou Secteurs
    data_type = st.radio("Afficher :", ["Indices", "Secteurs"], horizontal=True)

    if data_type == "Indices":
        selected = st.selectbox("Sélectionne un indice", list(INDICES.keys()))
        df = df_indices
        name = selected
    else:
        selected = st.selectbox("Sélectionne un secteur", list(SECTORS.keys()))
        df = df_sectors
        name = selected

    if name in df.columns:
        # Graphique 1: Prix + SMA
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.plot(df.index, df[name], label=name, color='blue')
        if f"{name}_SMA20" in df.columns:
            ax.plot(df.index, df[f"{name}_SMA20"], "--", label="SMA20", color='orange')
        if f"{name}_SMA50" in df.columns:
            ax.plot(df.index, df[f"{name}_SMA50"], "-.", label="SMA50", color='green')
        if f"{name}_SMA200" in df.columns:
            ax.plot(df.index, df[f"{name}_SMA200"], ":", label="SMA200", color='red')
        ax.set_title(f"{name} - Prix et Moyennes Mobiles")
        ax.legend()
        ax.grid()
        st.pyplot(fig)

        # Graphique 2: RSI
        if f"{name}_RSI" in df.columns:
            fig, ax = plt.subplots(figsize=(10, 3))
            ax.plot(df.index, df[f"{name}_RSI"], label="RSI(14)", color='purple')
            ax.axhline(70, color='red', linestyle='--', label="Surachat (70)")
            ax.axhline(30, color='green', linestyle='--', label="Survente (30)")
            ax.set_title(f"{name} - RSI(14)")
            ax.legend()
            ax.grid()
            ax.set_ylim(0, 100)
            st.pyplot(fig)

        # Graphique 3: MACD
        if f"{name}_MACD" in df.columns:
            fig, ax = plt.subplots(figsize=(10, 4))
            ax.plot(df.index, df[f"{name}_MACD"], label="MACD", color='blue')
            ax.plot(df.index, df[f"{name}_MACD_Signal"], label="Signal", color='orange')
            ax.bar(df.index, df[f"{name}_MACD_Hist"],
                   label="Histogramme", color=np.where(df[f"{name}_MACD_Hist"] > 0, 'g', 'r'), alpha=0.3)
            ax.set_title(f"{name} - MACD(12,26,9)")
            ax.legend()
            ax.grid()
            st.pyplot(fig)

        # Graphique 4: Bollinger Bands
        if f"{name}_BB_Upper" in df.columns:
            fig, ax = plt.subplots(figsize=(10, 4))
            ax.plot(df.index, df[name], label="Prix", color='blue')
            ax.plot(df.index, df[f"{name}_BB_Upper"], label="Bande Supérieure", color='green', linestyle='--')
            ax.plot(df.index, df[f"{name}_BB_Middle"], label="Bande Moyenne", color='orange')
            ax.plot(df.index, df[f"{name}_BB_Lower"], label="Bande Inférieure", color='red', linestyle='--')
            ax.set_title(f"{name} - Bollinger Bands(20,2)")
            ax.legend()
            ax.grid()
            st.pyplot(fig)

with tab3:
    st.subheader("Secteurs Bullish/Beash")
    bullish_sectors = trends_sectors[trends_sectors["> SMA20"] == "✅"].index.tolist()
    bearish_sectors = trends_sectors[trends_sectors["> SMA20"] == "❌"].index.tolist()

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**🟢 Secteurs BULLISH (au-dessus SMA20)**")
        for sector in bullish_sectors:
            st.markdown(f"- {sector}")
    with col2:
        st.markdown("**🔴 Secteurs BEARISH (en dessous SMA20)**")
        for sector in bearish_sectors:
            st.markdown(f"- {sector}")

    st.markdown("---")
    st.subheader("Signaux RSI")
    overbought = trends_sectors[trends_sectors["RSI_Surachat"] == "✅"].index.tolist()
    oversold = trends_sectors[trends_sectors["RSI_Survente"] == "✅"].index.tolist()

    col3, col4 = st.columns(2)
    with col3:
        st.markdown("**🔴 Surachat (RSI > 70)**")
        for sector in overbought:
            st.markdown(f"- {sector}")
    with col4:
        st.markdown("**🟢 Survente (RSI < 30)**")
        for sector in oversold:
            st.markdown(f"- {sector}")